"""
Admin router - company tracking and analytics
Admin-only endpoints for Luma HQ
"""

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, desc, and_
from datetime import datetime, date, timedelta
from typing import Optional, Literal
import logging
import csv
import io

from db import get_db
from models.company import Company
from models.document import Document, DocumentStatus
from models.record import Record
from models.usage_log import UsageLog
from models.company_stats import CompanyStats
from models.report import Report
from middleware import require_admin
from fastapi import Request

router = APIRouter(dependencies=[Depends(lambda request: require_admin(request))])
logger = logging.getLogger(__name__)


@router.get("/companies")
async def list_companies(
    db: Session = Depends(get_db),
    status_filter: Optional[str] = Query(None, description="Filter by status"),
    sector: Optional[str] = Query(None, description="Filter by sector")
):
    """
    List all companies with quick stats
    
    Returns: Company list with last 30 days uploads and 12 months emissions
    """
    try:
        # Base query
        query = db.query(Company)
        
        # Apply filters
        if status_filter:
            query = query.filter(Company.status == status_filter)
        if sector:
            query = query.filter(Company.sector.ilike(f"%{sector}%"))
        
        companies = query.order_by(Company.created_at.desc()).all()
        
        # Calculate stats for each company
        result = []
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        twelve_months_ago = datetime.utcnow() - timedelta(days=365)
        
        for company in companies:
            # Get document IDs for this company
            doc_ids = db.query(Document.id).filter(
                Document.company_id == company.id
            ).subquery()
            
            # Uploads in last 30 days
            uploads_30d = db.query(func.count(UsageLog.id)).filter(
                UsageLog.company_id == company.id,
                UsageLog.event_type == "upload",
                UsageLog.timestamp >= thirty_days_ago
            ).scalar() or 0
            
            # Total emissions last 12 months
            total_emissions_12m = db.query(func.sum(Record.co2e)).filter(
                Record.document_id.in_(doc_ids),
                Record.date >= twelve_months_ago.date()
            ).scalar() or 0
            
            result.append({
                **company.to_dict(),
                "uploads_30d": uploads_30d,
                "total_emissions_12m": float(total_emissions_12m)
            })
        
        return {
            "total": len(result),
            "companies": result
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to list companies: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/company/{company_id}")
async def get_company_detail(
    company_id: str,
    db: Session = Depends(get_db)
):
    """
    Get detailed company profile with full analytics
    """
    try:
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Company not found")
        
        # Get document IDs
        doc_ids = db.query(Document.id).filter(
            Document.company_id == company_id
        ).subquery()
        
        # Last 12 months emissions time series
        twelve_months_ago = datetime.utcnow() - timedelta(days=365)
        monthly_emissions = db.query(
            func.date_trunc('month', Record.date).label('month'),
            func.sum(Record.co2e).label('total')
        ).filter(
            Record.document_id.in_(doc_ids),
            Record.date >= twelve_months_ago.date()
        ).group_by(
            func.date_trunc('month', Record.date)
        ).order_by('month').all()
        
        timeseries = [
            {
                "month": month.strftime('%Y-%m') if month else None,
                "emissions": float(total or 0)
            }
            for month, total in monthly_emissions
        ]
        
        # Scope breakdown (all time)
        scope_totals = db.query(
            Record.scope,
            func.sum(Record.co2e).label('total')
        ).filter(
            Record.document_id.in_(doc_ids)
        ).group_by(Record.scope).all()
        
        scope_breakdown = {
            "scope1": 0,
            "scope2": 0,
            "scope3": 0
        }
        for scope, total in scope_totals:
            if scope:
                scope_breakdown[f"scope{scope}"] = float(total or 0)
        
        # Last 20 usage events
        usage_events = db.query(UsageLog).filter(
            UsageLog.company_id == company_id
        ).order_by(
            UsageLog.timestamp.desc()
        ).limit(20).all()
        
        # Recent documents
        documents = db.query(Document).filter(
            Document.company_id == company_id
        ).order_by(
            Document.upload_date.desc()
        ).limit(10).all()
        
        # Recent reports
        reports = db.query(Report).filter(
            Report.company_id == company_id
        ).order_by(
            Report.created_at.desc()
        ).limit(10).all()
        
        return {
            "company": company.to_dict(),
            "timeseries": timeseries,
            "scope_breakdown": scope_breakdown,
            "recent_events": [event.to_dict() for event in usage_events],
            "recent_documents": [doc.to_dict() for doc in documents],
            "recent_reports": [rep.to_dict() for rep in reports]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Failed to get company detail: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/activity")
async def get_activity_log(
    db: Session = Depends(get_db),
    event_type: Optional[str] = Query(None, description="Filter by event type"),
    company_id: Optional[str] = Query(None, description="Filter by company"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500)
):
    """
    Get paginated activity log
    """
    try:
        # Base query
        query = db.query(UsageLog).join(Company, UsageLog.company_id == Company.id)
        
        # Apply filters
        if event_type:
            query = query.filter(UsageLog.event_type == event_type)
        if company_id:
            query = query.filter(UsageLog.company_id == company_id)
        
        # Count total
        total = query.count()
        
        # Paginate
        offset = (page - 1) * page_size
        logs = query.order_by(
            UsageLog.timestamp.desc()
        ).limit(page_size).offset(offset).all()
        
        # Enrich with company names
        result = []
        for log in logs:
            log_dict = log.to_dict()
            log_dict['company_name'] = log.company.name if log.company else "Unknown"
            result.append(log_dict)
        
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": (total + page_size - 1) // page_size,
            "logs": result
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to get activity log: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/insights")
async def get_insights(
    db: Session = Depends(get_db),
    from_date: Optional[date] = Query(None, description="Start date"),
    to_date: Optional[date] = Query(None, description="End date")
):
    """
    Get aggregated insights and KPIs
    """
    try:
        # Default to last 12 months
        if not to_date:
            to_date = datetime.utcnow().date()
        if not from_date:
            from_date = to_date - timedelta(days=365)
        
        # Active companies (logged in or uploaded in last 30 days)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        active_companies = db.query(
            func.count(func.distinct(UsageLog.company_id))
        ).filter(
            UsageLog.timestamp >= thirty_days_ago
        ).scalar() or 0
        
        # Total emissions in period
        all_doc_ids = db.query(Document.id).subquery()
        total_emissions = db.query(func.sum(Record.co2e)).filter(
            Record.document_id.in_(all_doc_ids),
            Record.date.between(from_date, to_date)
        ).scalar() or 0
        
        # Reports generated in period
        reports_count = db.query(func.count(Report.id)).filter(
            Report.created_at >= datetime.combine(from_date, datetime.min.time()),
            Report.created_at <= datetime.combine(to_date, datetime.max.time())
        ).scalar() or 0
        
        # Upload success rate
        total_uploads = db.query(func.count(Document.id)).filter(
            Document.upload_date >= datetime.combine(from_date, datetime.min.time())
        ).scalar() or 1
        
        successful_uploads = db.query(func.count(Document.id)).filter(
            Document.upload_date >= datetime.combine(from_date, datetime.min.time()),
            Document.status == DocumentStatus.COMPLETED
        ).scalar() or 0
        
        upload_success_rate = successful_uploads / total_uploads if total_uploads > 0 else 0
        
        # Monthly timeseries
        monthly_data = db.query(
            func.date_trunc('month', Record.date).label('month'),
            func.sum(Record.co2e).label('emissions'),
            func.count(func.distinct(Record.document_id)).label('uploads'),
            func.count(func.distinct(Document.company_id)).label('active_companies')
        ).join(
            Document, Record.document_id == Document.id
        ).filter(
            Record.date.between(from_date, to_date)
        ).group_by(
            func.date_trunc('month', Record.date)
        ).order_by('month').all()
        
        timeseries = [
            {
                "month": month.strftime('%Y-%m') if month else None,
                "emissions": float(emissions or 0),
                "uploads": uploads or 0,
                "active_companies": active_companies or 0
            }
            for month, emissions, uploads, active_companies in monthly_data
        ]
        
        # Top companies by emissions
        top_companies = db.query(
            Company.id,
            Company.name,
            func.sum(Record.co2e).label('emissions_12m')
        ).join(
            Document, Document.company_id == Company.id
        ).join(
            Record, Record.document_id == Document.id
        ).filter(
            Record.date >= (datetime.utcnow() - timedelta(days=365)).date()
        ).group_by(
            Company.id, Company.name
        ).order_by(
            desc('emissions_12m')
        ).limit(10).all()
        
        top_companies_list = [
            {
                "company_id": str(comp_id),
                "name": name,
                "emissions_12m": float(emissions or 0)
            }
            for comp_id, name, emissions in top_companies
        ]
        
        # Top categories
        top_categories = db.query(
            Record.category,
            func.sum(Record.co2e).label('emissions_12m')
        ).filter(
            Record.date >= (datetime.utcnow() - timedelta(days=365)).date()
        ).group_by(
            Record.category
        ).order_by(
            desc('emissions_12m')
        ).limit(10).all()
        
        top_categories_list = [
            {
                "category": category,
                "emissions_12m": float(emissions or 0)
            }
            for category, emissions in top_categories
        ]
        
        return {
            "totals": {
                "active_companies": active_companies,
                "total_emissions_tco2e": float(total_emissions),
                "reports_generated": reports_count,
                "upload_success_rate": round(upload_success_rate, 3)
            },
            "timeseries": timeseries,
            "top_companies": top_companies_list,
            "top_categories": top_categories_list
        }
        
    except Exception as e:
        logger.error(f"❌ Failed to get insights: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
async def export_data(
    db: Session = Depends(get_db),
    format: Literal["csv", "xlsx"] = Query("csv", description="Export format"),
    range: Literal["last_month", "last_12m", "all"] = Query("last_12m", description="Date range")
):
    """
    Export aggregated company stats
    """
    try:
        # Determine date range
        if range == "last_month":
            from_date = (datetime.utcnow() - timedelta(days=30)).date()
        elif range == "last_12m":
            from_date = (datetime.utcnow() - timedelta(days=365)).date()
        else:  # all
            from_date = date(2020, 1, 1)
        
        # Get aggregated data per company
        companies_data = db.query(
            Company.id,
            Company.name,
            Company.sector,
            Company.country,
            func.sum(Record.co2e).label('total_emissions'),
            func.count(func.distinct(Document.id)).label('documents_count'),
            func.count(Record.id).label('records_count')
        ).outerjoin(
            Document, Document.company_id == Company.id
        ).outerjoin(
            Record, and_(
                Record.document_id == Document.id,
                Record.date >= from_date
            )
        ).group_by(
            Company.id, Company.name, Company.sector, Company.country
        ).order_by(
            Company.name
        ).all()
        
        if format == "csv":
            # Generate CSV
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Header
            writer.writerow([
                "Company ID",
                "Company Name",
                "Sector",
                "Country",
                "Total Emissions (tCO₂e)",
                "Documents",
                "Records"
            ])
            
            # Data rows
            for row in companies_data:
                writer.writerow([
                    str(row.id),
                    row.name,
                    row.sector or "",
                    row.country,
                    float(row.total_emissions or 0),
                    row.documents_count or 0,
                    row.records_count or 0
                ])
            
            csv_content = output.getvalue()
            output.close()
            
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=luma_admin_export_{datetime.now().strftime('%Y%m%d')}.csv"
                }
            )
        
        else:  # xlsx
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import tempfile
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Company Stats"
            
            # Header
            header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            headers = ["Company ID", "Company Name", "Sector", "Country", "Total Emissions (tCO₂e)", "Documents", "Records"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            # Data rows
            for row_num, data_row in enumerate(companies_data, start=2):
                ws.cell(row=row_num, column=1, value=str(data_row.id))
                ws.cell(row=row_num, column=2, value=data_row.name)
                ws.cell(row=row_num, column=3, value=data_row.sector or "")
                ws.cell(row=row_num, column=4, value=data_row.country)
                ws.cell(row=row_num, column=5, value=float(data_row.total_emissions or 0))
                ws.cell(row=row_num, column=6, value=data_row.documents_count or 0)
                ws.cell(row=row_num, column=7, value=data_row.records_count or 0)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to temp file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(temp_file.name)
            temp_file.close()
            
            with open(temp_file.name, "rb") as f:
                xlsx_content = f.read()
            
            os.unlink(temp_file.name)
            
            return Response(
                content=xlsx_content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=luma_admin_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
                }
            )
        
    except Exception as e:
        logger.error(f"❌ Export failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/aggregate-stats")
async def aggregate_monthly_stats(
    db: Session = Depends(get_db),
    month: Optional[date] = Query(None, description="Month to aggregate (YYYY-MM-01)")
):
    """
    Manually trigger monthly stats aggregation
    (In production, this would be a scheduled job)
    """
    try:
        if not month:
            # Default to last month
            today = datetime.utcnow().date()
            month = date(today.year, today.month, 1) - timedelta(days=1)
            month = date(month.year, month.month, 1)
        
        # Get all companies
        companies = db.query(Company).all()
        
        aggregated = 0
        for company in companies:
            # Calculate stats for this company and month
            month_start = month
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            # Count uploads
            uploads_count = db.query(func.count(UsageLog.id)).filter(
                UsageLog.company_id == company.id,
                UsageLog.event_type == "upload",
                UsageLog.timestamp >= datetime.combine(month_start, datetime.min.time()),
                UsageLog.timestamp < datetime.combine(month_end, datetime.max.time())
            ).scalar() or 0
            
            # Count records and emissions
            doc_ids = db.query(Document.id).filter(
                Document.company_id == company.id
            ).subquery()
            
            records_count = db.query(func.count(Record.id)).filter(
                Record.document_id.in_(doc_ids),
                Record.date >= month_start,
                Record.date <= month_end
            ).scalar() or 0
            
            total_emissions = db.query(func.sum(Record.co2e)).filter(
                Record.document_id.in_(doc_ids),
                Record.date >= month_start,
                Record.date <= month_end
            ).scalar() or 0
            
            # Reports generated
            reports_count = db.query(func.count(Report.id)).filter(
                Report.company_id == company.id,
                Report.created_at >= datetime.combine(month_start, datetime.min.time()),
                Report.created_at < datetime.combine(month_end, datetime.max.time())
            ).scalar() or 0
            
            # Upsert stats
            stats = db.query(CompanyStats).filter(
                CompanyStats.company_id == company.id,
                CompanyStats.month == month_start
            ).first()
            
            if stats:
                stats.uploads_count = uploads_count
                stats.records_count = records_count
                stats.total_emissions = total_emissions
                stats.reports_generated = reports_count
            else:
                stats = CompanyStats(
                    company_id=company.id,
                    month=month_start,
                    uploads_count=uploads_count,
                    records_count=records_count,
                    total_emissions=total_emissions,
                    reports_generated=reports_count,
                    active_users=1 if uploads_count > 0 else 0
                )
                db.add(stats)
            
            aggregated += 1
        
        db.commit()
        
        logger.info(f"✅ Aggregated stats for {aggregated} companies for {month}")
        
        return {
            "message": f"Successfully aggregated stats for {aggregated} companies",
            "month": month.isoformat(),
            "companies_processed": aggregated
        }
        
    except Exception as e:
        logger.error(f"❌ Stats aggregation failed: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
