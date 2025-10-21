"""
Report generation utilities - PDF and Excel exports
"""

import os
import logging
from datetime import datetime
from typing import Dict, List
import uuid

logger = logging.getLogger(__name__)

# Reports directory
REPORTS_DIR = os.getenv("REPORTS_DIR", "./reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


def generate_pdf_report(data: Dict) -> str:
    """
    Generate CSRD-Lite PDF report
    
    Uses ReportLab for PDF generation
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.charts.linecharts import HorizontalLineChart
        
        company = data['company']
        year = data['year']
        
        # Create PDF file
        filename = f"luma_csrd_report_{company.name.replace(' ', '_')}_{year}_{uuid.uuid4().hex[:8]}.pdf"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C5F2D'),  # Sage green
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2C5F2D'),
            spaceAfter=12
        )
        
        # Cover Page
        story.append(Spacer(1, 2*cm))
        story.append(Paragraph("CSRD Emission Report", title_style))
        story.append(Spacer(1, 1*cm))
        
        cover_data = [
            ["Company:", company.name],
            ["Sector:", company.sector or "Manufacturing"],
            ["Country:", company.country],
            ["Reporting Period:", f"{year}"],
            ["Report Date:", datetime.now().strftime("%d/%m/%Y")],
        ]
        
        cover_table = Table(cover_data, colWidths=[6*cm, 10*cm])
        cover_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2C5F2D')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(cover_table)
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        story.append(Spacer(1, 0.5*cm))
        
        summary_text = f"""
        This report presents the greenhouse gas (GHG) emissions inventory for {company.name} 
        for the calendar year {year}, prepared in alignment with the EU Corporate Sustainability 
        Reporting Directive (CSRD) and ESRS E1 - Climate Change standard.
        """
        story.append(Paragraph(summary_text, styles['BodyText']))
        story.append(Spacer(1, 0.5*cm))
        
        # Summary Table
        summary_data = [
            ["Metric", "Value", "Unit"],
            ["Total GHG Emissions", f"{data['total_co2e']:.2f}", "tonnes CO₂e"],
            ["Scope 1 (Direct)", f"{data['scope1']:.2f}", "tonnes CO₂e"],
            ["Scope 2 (Energy)", f"{data['scope2']:.2f}", "tonnes CO₂e"],
            ["Scope 3 (Indirect)", f"{data['scope3']:.2f}", "tonnes CO₂e"],
            ["Data Coverage", f"{data['coverage']:.1f}", "%"],
            ["Data Points", str(len(data['records'])), "records"],
        ]
        
        summary_table = Table(summary_data, colWidths=[8*cm, 4*cm, 4*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F2D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 1*cm))
        
        # Scope Breakdown (Pie Chart)
        story.append(Paragraph("Emissions by Scope", heading_style))
        story.append(Spacer(1, 0.3*cm))
        
        pie = Pie()
        pie.x = 200
        pie.y = 50
        pie.width = 150
        pie.height = 150
        pie.data = [data['scope1'], data['scope2'], data['scope3']]
        pie.labels = ['Scope 1', 'Scope 2', 'Scope 3']
        pie.slices.strokeWidth = 0.5
        pie.slices[0].fillColor = colors.HexColor('#8B4513')  # Brown
        pie.slices[1].fillColor = colors.HexColor('#DAA520')  # Gold
        pie.slices[2].fillColor = colors.HexColor('#2C5F2D')  # Green
        
        drawing = Drawing(400, 200)
        drawing.add(pie)
        story.append(drawing)
        story.append(Spacer(1, 1*cm))
        
        # Category Breakdown
        story.append(Paragraph("Top Emission Sources", heading_style))
        story.append(Spacer(1, 0.3*cm))
        
        breakdown = data['breakdown']
        top_sources = sorted(breakdown.items(), key=lambda x: x[1], reverse=True)[:5]
        
        sources_data = [["Category", "Emissions (tCO₂e)", "% of Total"]]
        for category, co2e in top_sources:
            percentage = (co2e / data['total_co2e'] * 100) if data['total_co2e'] > 0 else 0
            sources_data.append([
                category.replace('_', ' ').title(),
                f"{co2e:.2f}",
                f"{percentage:.1f}%"
            ])
        
        sources_table = Table(sources_data, colWidths=[8*cm, 4*cm, 4*cm])
        sources_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C5F2D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        
        story.append(sources_table)
        story.append(PageBreak())
        
        # Methodology
        story.append(Paragraph("Methodology", heading_style))
        story.append(Spacer(1, 0.3*cm))
        
        methodology_text = f"""
        <b>Calculation Approach:</b><br/>
        Emissions were calculated using internationally recognized emission factors from IPCC, EEA, DEFRA, and ADEME. 
        Activity data was extracted from utility bills, invoices, and operational records.<br/><br/>
        
        <b>Scope Definitions:</b><br/>
        • <b>Scope 1</b>: Direct emissions from sources owned or controlled by {company.name} (e.g., natural gas, company vehicles).<br/>
        • <b>Scope 2</b>: Indirect emissions from purchased electricity, steam, heating, and cooling.<br/>
        • <b>Scope 3</b>: Other indirect emissions from the value chain (e.g., purchased goods, logistics).<br/><br/>
        
        <b>Data Quality:</b><br/>
        {data['coverage']:.1f}% of identified emission sources have complete data. Gaps were addressed using 
        conservative estimation methods where necessary.<br/><br/>
        
        <b>Standards Compliance:</b><br/>
        This report aligns with ESRS E1 (Climate Change) disclosure requirements under the EU CSRD framework.
        """
        
        story.append(Paragraph(methodology_text, styles['BodyText']))
        story.append(Spacer(1, 1*cm))
        
        # Footer
        story.append(Spacer(1, 2*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"Generated by Luma Beta © {datetime.now().year}", footer_style))
        story.append(Paragraph("CSRD Automation Platform for EU SMEs", footer_style))
        
        # Build PDF
        doc.build(story)
        logger.info(f"✅ PDF report generated: {filepath}")
        
        return filepath
        
    except ImportError:
        logger.error("❌ reportlab not installed. Run: pip install reportlab")
        raise
    except Exception as e:
        logger.error(f"❌ PDF generation failed: {str(e)}")
        raise


def generate_excel_report(data: Dict) -> str:
    """
    Generate Excel datapoints report
    
    Uses openpyxl for Excel generation
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        company = data['company']
        year = data['year']
        
        # Create Excel file
        filename = f"luma_csrd_data_{company.name.replace(' ', '_')}_{year}_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        wb = Workbook()
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Header styling
        header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Summary data
        ws_summary['A1'] = "CSRD Emission Report"
        ws_summary['A1'].font = Font(bold=True, size=16, color="2C5F2D")
        ws_summary.merge_cells('A1:C1')
        
        ws_summary['A3'] = "Company"
        ws_summary['B3'] = company.name
        ws_summary['A4'] = "Year"
        ws_summary['B4'] = year
        ws_summary['A5'] = "Report Date"
        ws_summary['B5'] = datetime.now().strftime("%Y-%m-%d")
        
        ws_summary['A7'] = "Metric"
        ws_summary['B7'] = "Value"
        ws_summary['C7'] = "Unit"
        for cell in ['A7', 'B7', 'C7']:
            ws_summary[cell].fill = header_fill
            ws_summary[cell].font = header_font
        
        summary_rows = [
            ["Total Emissions", data['total_co2e'], "tonnes CO₂e"],
            ["Scope 1", data['scope1'], "tonnes CO₂e"],
            ["Scope 2", data['scope2'], "tonnes CO₂e"],
            ["Scope 3", data['scope3'], "tonnes CO₂e"],
            ["Data Coverage", data['coverage'], "%"],
        ]
        
        for i, row in enumerate(summary_rows, start=8):
            ws_summary[f'A{i}'] = row[0]
            ws_summary[f'B{i}'] = row[1]
            ws_summary[f'C{i}'] = row[2]
        
        # Sheet 2: Monthly Data
        ws_monthly = wb.create_sheet("Monthly Breakdown")
        
        ws_monthly['A1'] = "Month"
        ws_monthly['B1'] = "Emissions (tCO₂e)"
        for cell in ['A1', 'B1']:
            ws_monthly[cell].fill = header_fill
            ws_monthly[cell].font = header_font
        
        for i, month_data in enumerate(data['monthly_data'], start=2):
            ws_monthly[f'A{i}'] = month_data['month']
            ws_monthly[f'B{i}'] = month_data['co2e']
        
        # Sheet 3: Category Breakdown
        ws_category = wb.create_sheet("Category Breakdown")
        
        ws_category['A1'] = "Category"
        ws_category['B1'] = "Emissions (tCO₂e)"
        ws_category['C1'] = "% of Total"
        for cell in ['A1', 'B1', 'C1']:
            ws_category[cell].fill = header_fill
            ws_category[cell].font = header_font
        
        for i, (category, co2e) in enumerate(sorted(data['breakdown'].items(), key=lambda x: x[1], reverse=True), start=2):
            percentage = (co2e / data['total_co2e'] * 100) if data['total_co2e'] > 0 else 0
            ws_category[f'A{i}'] = category.replace('_', ' ').title()
            ws_category[f'B{i}'] = co2e
            ws_category[f'C{i}'] = f"{percentage:.2f}%"
        
        # Sheet 4: Detailed Records
        ws_records = wb.create_sheet("Detailed Records")
        
        headers = ["Date", "Supplier", "Category", "Usage", "Unit", "Cost (EUR)", "Scope", "CO₂e (tonnes)", "Factor Source"]
        for col, header in enumerate(headers, start=1):
            cell = ws_records.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for i, record in enumerate(data['records'], start=2):
            ws_records.cell(row=i, column=1, value=record.date.strftime("%Y-%m-%d") if record.date else "")
            ws_records.cell(row=i, column=2, value=record.supplier)
            ws_records.cell(row=i, column=3, value=record.category)
            ws_records.cell(row=i, column=4, value=float(record.usage) if record.usage else "")
            ws_records.cell(row=i, column=5, value=record.unit)
            ws_records.cell(row=i, column=6, value=float(record.cost) if record.cost else "")
            ws_records.cell(row=i, column=7, value=record.scope)
            ws_records.cell(row=i, column=8, value=float(record.co2e) if record.co2e else "")
            ws_records.cell(row=i, column=9, value=record.factor_source)
        
        # Adjust column widths
        for ws in [ws_summary, ws_monthly, ws_category, ws_records]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        logger.info(f"✅ Excel report generated: {filepath}")
        
        return filepath
        
    except ImportError:
        logger.error("❌ openpyxl not installed. Run: pip install openpyxl")
        raise
    except Exception as e:
        logger.error(f"❌ Excel generation failed: {str(e)}")
        raise
